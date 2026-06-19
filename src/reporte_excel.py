"""
reporte_excel.py
----------------
Reportes de facturas en Excel (openpyxl) para enviar por WhatsApp.

Reutiliza la lógica de consulta YA existente:
  - El mismo DataFrame `facturas_mensual` que el bot consulta hoy
    (columnas: folio, cliente, fecha, concepto, total, fecha_pago).
  - query_engine._mask_cliente para el match de cliente (substring + fuzzy),
    idéntico al que usa el bot al responder consultas de texto.

Aquí solo se: detecta la solicitud, filtra y construye el .xlsx. El envío por
WhatsApp (canal Meta) vive en webhook.py.
"""

from __future__ import annotations

import os
import re
import tempfile
from pathlib import Path

import pandas as pd

# Tope de filas: evita Excels gigantes y mantenerse holgado bajo el límite de
# tamaño de media de Meta. Reportes más grandes se acotan y se sugiere un mes.
MAX_FILAS = 1000

_MES_A_NUM = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}
_NUM_A_MES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre",
    12: "Diciembre",
}

# Palabras que NO forman parte del nombre del cliente al parsear la solicitud.
_RELLENO = {
    "reporte", "reportes", "reporta", "factura", "facturas", "de", "del", "la",
    "el", "los", "las", "en", "mensual", "semanal", "excel", "exportar",
    "exporta", "mandame", "mándame", "enviame", "envíame", "manda", "envia",
    "envía", "quiero", "dame", "porfa", "por", "favor", "un", "una",
}


def es_solicitud_dashboard(texto: str) -> bool:
    """True si el usuario pide el DASHBOARD HTML (con gráficas), no el Excel.

    Gatillos: "dashboard" en cualquier parte, o "reporte ..." acompañado de una
    palabra visual ("visual", "gráfico/a(s)", "grafico/a(s)"). Se usa para NO
    tratar estos mensajes como reporte Excel y enrutarlos al dashboard.
    """
    if not texto:
        return False
    t = texto.strip().lower()
    if "dashboard" in t:
        return True
    if t.startswith("reporte") and ("visual" in t or "gráfic" in t or "grafic" in t):
        return True
    return False


def parsear_solicitud_reporte(texto: str) -> dict | None:
    """Detecta una solicitud de reporte EXPORTABLE y extrae cliente y/o mes.

    Devuelve {"cliente": str|None, "mes": int|None, "mes_nombre": str|None}, o
    None si el texto no es una solicitud de reporte accionable.

    Casos:
      - "reporte de noviembre"            → mes=11
      - "facturas de Waldos de diciembre" → cliente="waldos", mes=12
      - "reporte" / "reporte mensual"     → mes = mes actual (reporte del mes)
      - "reporte semanal"                 → dias=7 (últimos 7 días)
      - "facturas" a secas                → None (lo maneja el flujo de texto)
    """
    if not texto:
        return None
    t = texto.strip().lower()
    if not (t.startswith("reporte") or t.startswith("factura")):
        return None
    # "reporte visual"/"gráfico"/"dashboard" piden el dashboard HTML, no el Excel.
    if es_solicitud_dashboard(texto):
        return None

    tokens = re.findall(r"[0-9a-záéíóúñü]+", t)

    mes_num = None
    mes_token = None
    for tok in tokens:
        if tok in _MES_A_NUM:
            mes_num = _MES_A_NUM[tok]
            mes_token = tok
            break

    cliente_toks = [
        tok for tok in tokens if tok not in _RELLENO and tok != mes_token
    ]
    cliente = " ".join(cliente_toks).strip() or None

    if mes_num is None and cliente is None:
        # Sin cliente ni mes explícito: solo "reporte/reporte mensual/semanal"
        # cuenta como reporte (general). "facturas" a secas NO (lo maneja texto).
        if not t.startswith("reporte"):
            return None
        if "semanal" in tokens:
            return {"cliente": None, "mes": None, "mes_nombre": None,
                    "dias": 7, "etiqueta": "última semana"}
        # "reporte" o "reporte mensual" → mes actual
        from datetime import datetime
        m = datetime.now().month
        return {"cliente": None, "mes": m, "mes_nombre": _NUM_A_MES[m],
                "dias": None, "etiqueta": None}

    return {
        "cliente": cliente,
        "mes": mes_num,
        "mes_nombre": _NUM_A_MES.get(mes_num) if mes_num else None,
        "dias": None,
        "etiqueta": None,
    }


def filtrar_facturas(
    facturas: pd.DataFrame,
    cliente: str | None = None,
    mes: int | None = None,
    dias: int | None = None,
) -> tuple[pd.DataFrame, bool]:
    """Filtra las facturas por cliente y/o periodo (mes de emisión o últimos N días).

    Reutiliza query_engine._mask_cliente (substring + fuzzy) para el cliente, el
    mismo matching que usa el bot en sus consultas. Devuelve (df_filtrado,
    truncado), acotando a MAX_FILAS si hace falta.
    """
    if facturas is None or facturas.empty:
        return pd.DataFrame(), False

    df = facturas
    if cliente:
        from src.query_engine import _mask_cliente
        df = df[_mask_cliente(df["cliente"], cliente.upper())]

    if "fecha" in df.columns:
        # La columna "fecha" puede llegar como string/object desde Postgres; hay
        # que normalizarla a datetime ANTES de usar el accesor .dt o comparar por
        # fecha (si no, falla con "Can only use .dt accessor with datetimelike
        # values"). Las fechas inválidas se vuelven NaT y quedan FUERA de los
        # filtros por mes/días (NaT no cumple ninguna comparación).
        df = df.copy()
        df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
        if dias is not None:
            from datetime import datetime, timedelta
            limite = datetime.now() - timedelta(days=dias)
            df = df[df["fecha"] >= limite]
        elif mes is not None:
            df = df[df["fecha"].dt.month == mes]
        df = df.sort_values("fecha")

    truncado = len(df) > MAX_FILAS
    if truncado:
        df = df.head(MAX_FILAS)
    return df, truncado


def _estado(fecha_pago) -> str:
    return "Pagada" if pd.notna(fecha_pago) else "Pendiente"


def generar_excel(
    facturas: pd.DataFrame,
    titulo: str,
    ruta: Path | None = None,
) -> Path:
    """Construye un .xlsx con encabezados, filas y totales. Devuelve la ruta.

    Columnas: Folio, Cliente, Monto, Fecha, Estado. Si no se pasa `ruta`, escribe
    a un archivo temporal (el llamador debe borrarlo tras enviarlo).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if ruta is None:
        fd, nombre = tempfile.mkstemp(prefix="reporte_", suffix=".xlsx")
        os.close(fd)
        ruta = Path(nombre)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    encabezados = ["Folio", "Cliente", "Monto", "Fecha", "Estado"]
    fila_enc = 3
    for col, nombre in enumerate(encabezados, start=1):
        ws.cell(row=fila_enc, column=col, value=nombre).font = Font(bold=True)

    total = 0.0
    r = fila_enc + 1
    for _, f in facturas.iterrows():
        folio = int(f["folio"]) if pd.notna(f["folio"]) else ""
        monto = float(f["total"]) if pd.notna(f["total"]) else 0.0
        # to_datetime defensivo: la fecha puede venir como string/object o NaT.
        fecha_dt = pd.to_datetime(f["fecha"], errors="coerce") if "fecha" in facturas.columns else None
        fecha = fecha_dt.strftime("%d/%m/%Y") if pd.notna(fecha_dt) else ""
        ws.cell(row=r, column=1, value=folio)
        ws.cell(row=r, column=2, value=str(f.get("cliente", "")))
        c_monto = ws.cell(row=r, column=3, value=monto)
        c_monto.number_format = "$#,##0.00"
        ws.cell(row=r, column=4, value=fecha)
        ws.cell(row=r, column=5, value=_estado(f.get("fecha_pago")))
        total += monto
        r += 1

    ws.cell(row=r, column=2, value="TOTAL").font = Font(bold=True)
    c_tot = ws.cell(row=r, column=3, value=total)
    c_tot.font = Font(bold=True)
    c_tot.number_format = "$#,##0.00"
    ws.cell(row=r, column=5, value=f"{len(facturas)} facturas").font = Font(bold=True)

    for col, ancho in zip("ABCDE", (10, 32, 16, 14, 12)):
        ws.column_dimensions[col].width = ancho

    wb.save(ruta)
    return ruta
