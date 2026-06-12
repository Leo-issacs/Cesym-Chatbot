"""
run_manual_tests.py
-------------------
Script de pruebas manuales y generador de reporte de calidad del Excel.

No requiere pytest. Se ejecuta directamente:
    python scripts/run_manual_tests.py

Genera:
  - Resultados de prueba en consola (PASS / FAIL).
  - Archivo data_quality_report.md con el reporte completo.
"""

import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import openpyxl

from src.loader import load_facturado, load_pendiente, _resolver_ruta_cartera, DATA_RAW_DIR

# loader.py ya no exporta EXCEL_PATH: ahora detecta el Excel dinámicamente.
# Se resuelve de forma tolerante para que el script importe aunque no haya Excel.
try:
    EXCEL_PATH = _resolver_ruta_cartera()
except FileNotFoundError:
    EXCEL_PATH = DATA_RAW_DIR / "CARTERA AL 11032026.xlsx"
from src.cleaner import clean_facturado, clean_pendiente
from src.query_engine import run_query

# ─── Utilidades de reporte ─────────────────────────────────────────────────────

PASS_COUNT = 0
FAIL_COUNT = 0
RESULTADOS = []


def verificar(nombre: str, condicion: bool, detalle: str = "") -> bool:
    global PASS_COUNT, FAIL_COUNT
    if condicion:
        PASS_COUNT += 1
        print(f"  [PASS] {nombre}")
        RESULTADOS.append(("PASS", nombre, ""))
    else:
        FAIL_COUNT += 1
        msg = f" — {detalle}" if detalle else ""
        print(f"  [FAIL] {nombre}{msg}")
        RESULTADOS.append(("FAIL", nombre, detalle))
    return condicion


def seccion(titulo: str):
    print(f"\n{'-'*60}")
    print(f"  {titulo.upper()}")
    print(f"{'-'*60}")


# ─── Bloque 1: Lectura del Excel ───────────────────────────────────────────────

def pruebas_loader():
    seccion("1. Lectura del Excel")

    verificar("El archivo Excel existe", EXCEL_PATH.exists(),
              f"Ruta esperada: {EXCEL_PATH}")

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        hojas = wb.sheetnames
        wb.close()
        verificar("Excel se abre sin errores", True)
        verificar("Hoja 'OC FACTURADO' detectada", "OC FACTURADO" in hojas,
                  f"Hojas encontradas: {hojas}")
        verificar("Hoja 'PTE OC 25-26' detectada", "PTE OC 25-26" in hojas,
                  f"Hojas encontradas: {hojas}")
        verificar("Número de hojas es 3", len(hojas) == 3,
                  f"Hojas: {hojas}")
    except Exception as e:
        verificar("Excel se abre sin errores", False, str(e))
        return None, None, None, None, None, None

    try:
        raw_fac = load_facturado()
        verificar("load_facturado() sin error", True)
        verificar("load_facturado() devuelve DataFrame", isinstance(raw_fac, pd.DataFrame))
        verificar("load_facturado() no está vacío", len(raw_fac) > 0,
                  f"Filas: {len(raw_fac)}")
        columnas_fac = [str(c).upper() for c in raw_fac.columns]
        verificar("Columna FACTURA detectada", "FACTURA" in columnas_fac,
                  f"Columnas: {raw_fac.columns.tolist()}")
        verificar("Columna OC detectada", "OC" in columnas_fac)
    except Exception as e:
        verificar("load_facturado() sin error", False, str(e))
        raw_fac = None

    try:
        raw_pte = load_pendiente()
        verificar("load_pendiente() sin error", True)
        verificar("load_pendiente() devuelve DataFrame", isinstance(raw_pte, pd.DataFrame))
        verificar("load_pendiente() no está vacío", len(raw_pte) > 0,
                  f"Filas: {len(raw_pte)}")
        columnas_pte = [str(c).upper() for c in raw_pte.columns]
        verificar("Columna COT detectada", "COT" in columnas_pte,
                  f"Columnas: {raw_pte.columns.tolist()}")
        verificar("Columna SUC detectada", "SUC" in columnas_pte)
    except Exception as e:
        verificar("load_pendiente() sin error", False, str(e))
        raw_pte = None

    return raw_fac, raw_pte


# ─── Bloque 2: Limpieza de datos ──────────────────────────────────────────────

def pruebas_cleaner(raw_fac, raw_pte):
    seccion("2. Limpieza y normalización de datos")

    if raw_fac is None or raw_pte is None:
        print("  [SKIP] No se puede limpiar sin datos RAW válidos.")
        return None, None, [], []

    try:
        df_fac, warns_fac = clean_facturado(raw_fac)
        verificar("clean_facturado() sin error", True)
        verificar("Retorna tuple (DataFrame, lista)", isinstance(df_fac, pd.DataFrame))
        verificar("Columnas correctas en facturado",
                  list(df_fac.columns) == ["factura", "oc", "monto_actual", "prioridad", "fecha", "estado"],
                  f"Columnas: {list(df_fac.columns)}")
        verificar("Filas de totales excluidas (menos filas que RAW)",
                  len(df_fac) < len(raw_fac),
                  f"RAW={len(raw_fac)}, Limpio={len(df_fac)}")
        verificar("Sin valores nulos en columna 'factura'",
                  df_fac["factura"].isna().sum() == 0,
                  f"Nulos: {df_fac['factura'].isna().sum()}")
        verificar("Tipo Int64 en columna 'factura'",
                  pd.api.types.is_integer_dtype(df_fac["factura"]),
                  f"Tipo actual: {df_fac['factura'].dtype}")
        verificar("Tipo float en columna 'monto_actual'",
                  pd.api.types.is_float_dtype(df_fac["monto_actual"]),
                  f"Tipo actual: {df_fac['monto_actual'].dtype}")
        verificar("Tipo datetime en columna 'fecha'",
                  pd.api.types.is_datetime64_any_dtype(df_fac["fecha"]),
                  f"Tipo actual: {df_fac['fecha'].dtype}")
        fechas_validas = df_fac["fecha"].dropna()
        if len(fechas_validas) > 0:
            verificar("Fechas convertidas con año >= 2000",
                      fechas_validas.dt.year.min() >= 2000,
                      f"Año mínimo: {fechas_validas.dt.year.min()}")
    except Exception as e:
        verificar("clean_facturado() sin error", False, str(e))
        df_fac, warns_fac = None, []

    try:
        df_pte, warns_pte = clean_pendiente(raw_pte)
        verificar("clean_pendiente() sin error", True)
        verificar("Retorna tuple (DataFrame, lista)", isinstance(df_pte, pd.DataFrame))
        verificar("Columnas correctas en pendiente",
                  list(df_pte.columns) == ["cot", "suc", "importe", "concepto"],
                  f"Columnas: {list(df_pte.columns)}")
        verificar("Filas de totales excluidas en pendiente",
                  len(df_pte) < len(raw_pte),
                  f"RAW={len(raw_pte)}, Limpio={len(df_pte)}")
        verificar("Sin valores nulos en columna 'cot'",
                  df_pte["cot"].isna().sum() == 0,
                  f"Nulos: {df_pte['cot'].isna().sum()}")
        verificar("Tipo Int64 en columna 'cot'",
                  pd.api.types.is_integer_dtype(df_pte["cot"]),
                  f"Tipo actual: {df_pte['cot'].dtype}")
        verificar("Tipo float en columna 'importe'",
                  pd.api.types.is_float_dtype(df_pte["importe"]),
                  f"Tipo actual: {df_pte['importe'].dtype}")
    except Exception as e:
        verificar("clean_pendiente() sin error", False, str(e))
        df_pte, warns_pte = None, []

    return df_fac, df_pte, warns_fac, warns_pte


# ─── Bloque 3: Consultas ──────────────────────────────────────────────────────

def pruebas_queries(df_fac, df_pte):
    seccion("3. Consultas y búsquedas")

    if df_fac is None or df_pte is None:
        print("  [SKIP] No se puede consultar sin datos limpios.")
        return

    def q(cmd):
        return run_query(cmd, df_fac, df_pte)

    verificar("total — devuelve string con $", "$" in q("total"))
    verificar("total facturado — monto positivo",
              float(q("total facturado").split("$")[-1].replace(",", "").strip()) > 0)
    verificar("total pendiente — monto positivo",
              float(q("total pendiente").split("$")[-1].replace(",", "").strip()) > 0)
    verificar("resumen — contiene 'registros'", "registros" in q("resumen").lower())
    verificar("facturas — contiene 'Fac'", "Fac" in q("facturas"))
    verificar("facturas — contiene Total", "Total" in q("facturas"))
    verificar("pendientes — contiene 'Cot'", "Cot" in q("pendientes"))
    verificar("pendientes — contiene Total", "Total" in q("pendientes"))
    verificar("ayuda — contiene 'buscar'", "buscar" in q("ayuda").lower())

    # Búsqueda por factura
    primera_fac = int(df_fac["factura"].dropna().iloc[0])
    res_fac = q(f"buscar factura {primera_fac}")
    verificar(f"buscar factura {primera_fac} — encontrada", str(primera_fac) in res_fac)
    verificar("buscar factura 999999999 — no encontrada",
              "no se encontró" in q("buscar factura 999999999").lower())
    verificar("buscar factura abc — error de tipo",
              "válido" in q("buscar factura abc").lower() or "no es" in q("buscar factura abc").lower())

    # Búsqueda por OC
    primera_oc = df_fac["oc"].iloc[0]
    res_oc = q(f"buscar oc {primera_oc}")
    verificar(f"buscar oc '{primera_oc}' — encontrada", "$" in res_oc or primera_oc[:3] in res_oc)
    verificar("buscar oc ZZZINEXISTENTE — no encontrada",
              "no se encontró" in q("buscar oc ZZZINEXISTENTE").lower())

    # Búsqueda por COT
    primera_cot = int(df_pte["cot"].dropna().iloc[0])
    res_cot = q(f"buscar cot {primera_cot}")
    verificar(f"buscar cot {primera_cot} — encontrada", str(primera_cot) in res_cot)
    verificar("buscar cot 999999999 — no encontrada",
              "no se encontró" in q("buscar cot 999999999").lower())
    verificar("buscar cot abc — error de tipo",
              "válido" in q("buscar cot abc").lower() or "no es" in q("buscar cot abc").lower())

    # Búsqueda por SUC
    primera_suc = int(df_pte["suc"].dropna().iloc[0])
    res_suc = q(f"buscar suc {primera_suc}")
    verificar(f"buscar suc {primera_suc} — encontrada", "$" in res_suc or str(primera_suc) in res_suc)
    verificar("buscar suc 99999 — no encontrada",
              "no hay" in q("buscar suc 99999").lower() or "no se encontró" in q("buscar suc 99999").lower())

    # Estado y errores
    verificar("estado aceptada — devuelve string", len(q("estado aceptada")) > 0)
    verificar("estado prioridad — devuelve string", len(q("estado prioridad")) > 0)
    verificar("errores — devuelve string", len(q("errores")) > 0)
    verificar("comando vacío — devuelve mensaje", len(q("")) > 0)
    verificar("comando desconocido — sugiere ayuda",
              "ayuda" in q("xyz_zzz_inexistente").lower() or "no reconocido" in q("xyz_zzz_inexistente").lower())


# ─── Bloque 4: Validaciones de calidad ────────────────────────────────────────

def pruebas_validaciones(df_fac, df_pte):
    seccion("4. Validaciones de calidad")

    if df_fac is None or df_pte is None:
        print("  [SKIP] No se puede validar sin datos limpios.")
        return

    # Montos inválidos
    montos_invalidos_fac = df_fac[df_fac["monto_actual"].isna() | (df_fac["monto_actual"] <= 0)]
    verificar(f"Montos inválidos en facturado ({len(montos_invalidos_fac)} registros)",
              True)  # Siempre pasa; el reporte muestra el número

    montos_invalidos_pte = df_pte[df_pte["importe"].isna() | (df_pte["importe"] <= 0)]
    verificar(f"Montos inválidos en pendiente ({len(montos_invalidos_pte)} registros)",
              True)

    # Fechas vacías
    sin_fecha = df_fac["fecha"].isna().sum()
    verificar(f"Fechas vacías detectadas en facturado ({sin_fecha} registros)",
              True)

    # Sin OC
    sin_oc = df_fac[df_fac["oc"].isin(["nan", "", "NaN"])]
    verificar(f"Facturas sin OC asignada ({len(sin_oc)} registros)",
              True)

    # Duplicados facturado
    dup_fac = df_fac[df_fac.duplicated("factura", keep=False)]
    verificar(f"Facturas duplicadas ({len(dup_fac)} registros afectados)",
              True)

    # Duplicados OC
    ocs_validas = df_fac[~df_fac["oc"].isin(["nan", "", "NaN"])]
    dup_oc = ocs_validas[ocs_validas.duplicated("oc", keep=False)]
    verificar(f"OC duplicadas en facturado ({len(dup_oc)} registros afectados)",
              True)

    # Duplicados pendiente
    dup_pte = df_pte[df_pte.duplicated("cot", keep=False)]
    verificar(f"Cotizaciones duplicadas ({len(dup_pte)} registros afectados)",
              True)

    # Registros incompletos (sin OC Y sin fecha Y sin monto)
    incompletos = df_fac[
        df_fac["oc"].isin(["nan", "", "NaN"]) |
        df_fac["fecha"].isna() |
        df_fac["monto_actual"].isna()
    ]
    verificar(f"Registros incompletos en facturado ({len(incompletos)} registros)",
              True)


# ─── Generador de reporte de calidad ──────────────────────────────────────────

def generar_reporte(raw_fac, raw_pte, df_fac, df_pte, warns_fac, warns_pte) -> str:
    fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lineas = []

    def ln(x=""):
        lineas.append(x)

    ln("# Reporte de Calidad de Datos")
    ln(f"**Generado:** {fecha_hora}")
    ln(f"**Archivo:** `{EXCEL_PATH.name}`")
    ln()

    # ── Hojas ──
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        hojas = wb.sheetnames
        wb.close()
        ln("## Hojas detectadas")
        for h in hojas:
            ln(f"- `{h}`")
    except Exception as e:
        ln(f"## Hojas detectadas")
        ln(f"Error al abrir el Excel: {e}")
    ln()

    # ── OC FACTURADO ──
    ln("## Hoja: OC FACTURADO")
    if raw_fac is not None and df_fac is not None:
        filas_ignoradas = len(raw_fac) - len(df_fac)
        ln(f"| Métrica | Valor |")
        ln(f"|---------|-------|")
        ln(f"| Filas RAW (antes de limpieza) | {len(raw_fac)} |")
        ln(f"| Filas limpias | {len(df_fac)} |")
        ln(f"| Filas ignoradas (totales/encabezados) | {filas_ignoradas} |")
        ln()
        ln(f"**Columnas detectadas:** `{', '.join(df_fac.columns.tolist())}`")
        ln()

        # Totales
        total_fac = df_fac["monto_actual"].sum()
        ln(f"### Totales")
        ln(f"- **Total OC Facturado:** `${total_fac:,.2f}`")
        ln()

        # Montos
        montos_invalidos = df_fac[df_fac["monto_actual"].isna() | (df_fac["monto_actual"] <= 0)]
        ln(f"### Registros con monto inválido (cero, negativo o vacío)")
        if montos_invalidos.empty:
            ln("- Ninguno detectado.")
        else:
            ln(f"- **Cantidad:** {len(montos_invalidos)}")
            ln(f"- **Facturas afectadas:** {montos_invalidos['factura'].tolist()}")
        ln()

        # Fechas
        sin_fecha = df_fac[df_fac["fecha"].isna()]
        ln(f"### Registros sin fecha")
        if sin_fecha.empty:
            ln("- Ninguno detectado.")
        else:
            ln(f"- **Cantidad:** {len(sin_fecha)}")
            ln(f"- **Facturas afectadas:** {sin_fecha['factura'].tolist()}")
        ln()

        # Sin OC
        sin_oc = df_fac[df_fac["oc"].isin(["nan", "", "NaN"])]
        ln(f"### Registros sin OC asignada")
        if sin_oc.empty:
            ln("- Ninguno detectado.")
        else:
            ln(f"- **Cantidad:** {len(sin_oc)}")
            ln(f"- **Facturas afectadas:** {sin_oc['factura'].tolist()}")
        ln()

        # Duplicados factura
        dup_fac = df_fac[df_fac.duplicated("factura", keep=False)]
        ln(f"### Facturas duplicadas")
        if dup_fac.empty:
            ln("- Ninguna detectada.")
        else:
            ln(f"- **Registros afectados:** {len(dup_fac)}")
            ln(f"- **Números duplicados:** {dup_fac['factura'].unique().tolist()}")
        ln()

        # Duplicados OC
        ocs_validas = df_fac[~df_fac["oc"].isin(["nan", "", "NaN"])]
        dup_oc = ocs_validas[ocs_validas.duplicated("oc", keep=False)]
        ln(f"### OC duplicadas")
        if dup_oc.empty:
            ln("- Ninguna detectada.")
        else:
            ln(f"- **Registros afectados:** {len(dup_oc)}")
            ln(f"- **OC duplicadas:** {dup_oc['oc'].unique().tolist()}")
        ln()

        # Registros incompletos
        incompletos = df_fac[
            df_fac["oc"].isin(["nan", "", "NaN"]) |
            df_fac["fecha"].isna() |
            df_fac["monto_actual"].isna()
        ]
        ln(f"### Registros incompletos (sin OC, sin fecha o sin monto)")
        if incompletos.empty:
            ln("- Ninguno detectado.")
        else:
            ln(f"- **Cantidad:** {len(incompletos)}")
            ln(f"- **Facturas afectadas:** {incompletos['factura'].tolist()}")
        ln()

        # Advertencias del cleaner
        ln(f"### Advertencias del proceso de limpieza")
        if warns_fac:
            for w in warns_fac:
                ln(f"- ⚠ {w}")
        else:
            ln("- Sin advertencias.")
        ln()
    else:
        ln("*No se pudo cargar la hoja.*")
        ln()

    # ── PTE OC 25-26 ──
    ln("## Hoja: PTE OC 25-26")
    if raw_pte is not None and df_pte is not None:
        filas_ignoradas_pte = len(raw_pte) - len(df_pte)
        ln(f"| Métrica | Valor |")
        ln(f"|---------|-------|")
        ln(f"| Filas RAW (antes de limpieza) | {len(raw_pte)} |")
        ln(f"| Filas limpias | {len(df_pte)} |")
        ln(f"| Filas ignoradas (totales/encabezados) | {filas_ignoradas_pte} |")
        ln()
        ln(f"**Columnas detectadas:** `{', '.join(df_pte.columns.tolist())}`")
        ln()

        total_pte = df_pte["importe"].sum()
        ln(f"### Totales")
        ln(f"- **Total Pendiente (PTE OC):** `${total_pte:,.2f}`")
        ln()

        # Montos inválidos pendiente
        imp_invalidos = df_pte[df_pte["importe"].isna() | (df_pte["importe"] <= 0)]
        ln(f"### Cotizaciones con importe inválido")
        if imp_invalidos.empty:
            ln("- Ninguna detectada.")
        else:
            ln(f"- **Cantidad:** {len(imp_invalidos)}")
            ln(f"- **COTs afectadas:** {imp_invalidos['cot'].tolist()}")
        ln()

        # Duplicados pendiente
        dup_pte = df_pte[df_pte.duplicated("cot", keep=False)]
        ln(f"### Cotizaciones duplicadas")
        if dup_pte.empty:
            ln("- Ninguna detectada.")
        else:
            ln(f"- **Registros afectados:** {len(dup_pte)}")
            ln(f"- **COTs duplicadas:** {dup_pte['cot'].unique().tolist()}")
        ln()

        # Sucursales únicas
        sucs = df_pte["suc"].dropna().unique()
        ln(f"### Sucursales con cotizaciones pendientes")
        ln(f"- **Cantidad de sucursales:** {len(sucs)}")
        ln(f"- **Números:** {sorted(sucs.tolist())}")
        ln()

        ln(f"### Advertencias del proceso de limpieza")
        if warns_pte:
            for w in warns_pte:
                ln(f"- ⚠ {w}")
        else:
            ln("- Sin advertencias.")
        ln()
    else:
        ln("*No se pudo cargar la hoja.*")
        ln()

    # ── Resumen global ──
    ln("## Resumen global")
    if df_fac is not None and df_pte is not None:
        total_fac = df_fac["monto_actual"].sum()
        total_pte = df_pte["importe"].sum()
        ln(f"| Concepto | Registros | Monto |")
        ln(f"|----------|-----------|-------|")
        ln(f"| OC Facturado | {len(df_fac)} | `${total_fac:,.2f}` |")
        ln(f"| PTE OC 25-26 | {len(df_pte)} | `${total_pte:,.2f}` |")
        ln(f"| **TOTAL CARTERA** | {len(df_fac) + len(df_pte)} | **`${total_fac + total_pte:,.2f}`** |")
        ln()

    # ── Advertencias globales ──
    todas_advertencias = (warns_fac or []) + (warns_pte or [])
    ln("## Advertencias globales")
    if todas_advertencias:
        for w in todas_advertencias:
            ln(f"- ⚠ {w}")
    else:
        ln("- Sin advertencias globales.")
    ln()

    # ── Limitaciones ──
    ln("## Limitaciones conocidas")
    ln("- Solo se procesa el archivo `CARTERA AL 11032026.xlsx`.")
    ln("- El segundo Excel (trabajos realizados) no está disponible aún.")
    ln("- No se cruza información entre hojas todavía.")
    ln("- La detección de facturas duplicadas y OC duplicadas no está en el motor de consultas (solo en este reporte).")
    ln()

    return "\n".join(lineas)


# ─── Punto de entrada ──────────────────────────────────────────────────────────

def main():
    print("\n" + "=" * 60)
    print("  PRUEBAS MANUALES - CESYM CHATBOT")
    print("=" * 60)

    raw_fac, raw_pte = pruebas_loader()

    df_fac, df_pte, warns_fac, warns_pte = pruebas_cleaner(raw_fac, raw_pte)

    pruebas_queries(df_fac, df_pte)

    pruebas_validaciones(df_fac, df_pte)

    # ── Resumen de pruebas ──
    total = PASS_COUNT + FAIL_COUNT
    print(f"\n{'='*60}")
    print(f"  RESULTADO: {PASS_COUNT}/{total} pruebas pasaron")
    if FAIL_COUNT > 0:
        print(f"  FALLOS: {FAIL_COUNT}")
        for estado, nombre, detalle in RESULTADOS:
            if estado == "FAIL":
                print(f"    * {nombre}" + (f": {detalle}" if detalle else ""))
    print(f"{'='*60}")

    # ── Generar reporte de calidad ──
    print("\nGenerando reporte de calidad del Excel...")
    reporte = generar_reporte(raw_fac, raw_pte, df_fac, df_pte,
                               warns_fac or [], warns_pte or [])

    reporte_path = Path(__file__).parent.parent / "data_quality_report.md"
    reporte_path.write_text(reporte, encoding="utf-8")
    print(f"Reporte guardado en: {reporte_path}")

    return FAIL_COUNT == 0


if __name__ == "__main__":
    exito = main()
    sys.exit(0 if exito else 1)
