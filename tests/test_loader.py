"""
test_loader.py
--------------
Pruebas de lectura del Excel sin modificar el archivo original.

Valida:
  - Que el archivo Excel existe en la ruta esperada.
  - Que las hojas se detectan correctamente.
  - Que load_facturado() y load_pendiente() devuelven DataFrames válidos.
  - Que los encabezados dinámicos se detectan correctamente.
  - Que los DataFrames no están vacíos.
"""

import pandas as pd
import pytest
import openpyxl

from src.loader import load_facturado, load_pendiente, _resolver_ruta_cartera, DATA_RAW_DIR

# loader.py ya no exporta EXCEL_PATH: ahora detecta el Excel dinámicamente.
# Se resuelve aquí de forma tolerante para que pytest pueda RECOLECTAR este módulo
# aunque no haya Excel en data/raw/. Los tests de contenido fallarán por su cuenta
# (sin romper la recolección).
try:
    EXCEL_PATH = _resolver_ruta_cartera()
except FileNotFoundError:
    EXCEL_PATH = DATA_RAW_DIR / "CARTERA AL 11032026.xlsx"


class TestArchivoExcel:
    def test_excel_existe(self):
        assert EXCEL_PATH.exists(), f"No se encontró el Excel en: {EXCEL_PATH}"

    def test_excel_tiene_hoja_oc_facturado(self):
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        assert "OC FACTURADO" in wb.sheetnames, "Falta la hoja 'OC FACTURADO'"
        wb.close()

    def test_excel_tiene_hoja_pte_oc(self):
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        assert "PTE OC 25-26" in wb.sheetnames, "Falta la hoja 'PTE OC 25-26'"
        wb.close()

    def test_excel_tiene_exactamente_tres_hojas(self):
        """El Excel conocido tiene OC FACTURADO, PTE OC 25-26 y Hoja1."""
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
        assert len(wb.sheetnames) == 3, f"Número de hojas inesperado: {wb.sheetnames}"
        wb.close()


class TestLoadFacturado:
    def test_devuelve_dataframe(self, raw_facturado):
        assert isinstance(raw_facturado, pd.DataFrame)

    def test_no_esta_vacio(self, raw_facturado):
        assert len(raw_facturado) > 0, "load_facturado() devolvió un DataFrame vacío"

    def test_encabezado_detectado(self, raw_facturado):
        """El header dinámico debe haber encontrado la columna FACTURA."""
        columnas = [str(c).strip().upper() for c in raw_facturado.columns]
        assert "FACTURA" in columnas, f"Columna FACTURA no detectada. Columnas: {columnas}"

    def test_tiene_columna_oc(self, raw_facturado):
        columnas = [str(c).strip().upper() for c in raw_facturado.columns]
        assert "OC" in columnas, f"Columna OC no detectada. Columnas: {columnas}"

    def test_multiples_columnas(self, raw_facturado):
        assert raw_facturado.shape[1] >= 5, (
            f"Se esperaban al menos 5 columnas, se encontraron {raw_facturado.shape[1]}"
        )


class TestLoadPendiente:
    def test_devuelve_dataframe(self, raw_pendiente):
        assert isinstance(raw_pendiente, pd.DataFrame)

    def test_no_esta_vacio(self, raw_pendiente):
        assert len(raw_pendiente) > 0, "load_pendiente() devolvió un DataFrame vacío"

    def test_encabezado_detectado(self, raw_pendiente):
        """El header dinámico debe haber encontrado la columna COT."""
        columnas = [str(c).strip().upper() for c in raw_pendiente.columns]
        assert "COT" in columnas, f"Columna COT no detectada. Columnas: {columnas}"

    def test_tiene_columna_suc(self, raw_pendiente):
        columnas = [str(c).strip().upper() for c in raw_pendiente.columns]
        assert "SUC" in columnas, f"Columna SUC no detectada. Columnas: {columnas}"

    def test_multiples_columnas(self, raw_pendiente):
        assert raw_pendiente.shape[1] >= 4, (
            f"Se esperaban al menos 4 columnas, se encontraron {raw_pendiente.shape[1]}"
        )
