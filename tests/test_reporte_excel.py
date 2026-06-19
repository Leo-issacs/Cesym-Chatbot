"""
test_reporte_excel.py
---------------------
Reportes de facturas en Excel enviados por WhatsApp (canal Meta). Cubre:
  - Detección de la solicitud (cliente/periodo, "reporte mensual", "semanal").
  - Filtrado reutilizando la lógica de consulta (cliente/mes/días) + acotado.
  - Generación del .xlsx (encabezados, filas, totales).
  - Envío como documento Meta (subida de media + type document), httpx mockeado.
  - Intercepción SOLO en Meta: el camino Twilio queda intacto (sigue en texto).
"""

import asyncio

import pandas as pd
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import src.reporte_excel as rx
import src.webhook as webhook


def _facturas_demo():
    """DataFrame con las columnas reales de facturas_mensual."""
    return pd.DataFrame({
        "folio":      [101, 102, 103, 104],
        "cliente":    ["WALDOS", "WALDOS", "OXXO", "SORIANA"],
        "fecha":      pd.to_datetime(["2025-12-03", "2025-11-20", "2025-12-15", "2025-12-28"]),
        "concepto":   ["Mantenimiento", "Instalación", "Servicio", "Reparación"],
        "total":      [1500.0, 2300.0, 800.0, 1200.0],
        "fecha_pago": pd.to_datetime(["2025-12-10", pd.NaT, "2025-12-20", pd.NaT]),
    })


# ─── Detección de la solicitud ───────────────────────────────────────────────

def test_parsea_cliente_y_mes():
    s = rx.parsear_solicitud_reporte("facturas de Waldos de diciembre")
    assert s["cliente"] == "waldos"
    assert s["mes"] == 12 and s["mes_nombre"] == "Diciembre"


def test_parsea_solo_mes():
    s = rx.parsear_solicitud_reporte("reporte de noviembre")
    assert s["cliente"] is None and s["mes"] == 11


def test_reporte_mensual_usa_mes_actual():
    from datetime import datetime
    s = rx.parsear_solicitud_reporte("reporte mensual")
    assert s is not None and s["cliente"] is None
    assert s["mes"] == datetime.now().month


def test_reporte_semanal_usa_dias():
    s = rx.parsear_solicitud_reporte("reporte semanal")
    assert s is not None and s["dias"] == 7


def test_facturas_a_secas_no_es_reporte():
    assert rx.parsear_solicitud_reporte("facturas") is None
    assert rx.parsear_solicitud_reporte("total general") is None


# ─── Filtrado (reutiliza _mask_cliente) ──────────────────────────────────────

def test_filtra_por_cliente_y_mes():
    df, trunc = rx.filtrar_facturas(_facturas_demo(), cliente="waldos", mes=12)
    assert list(df["folio"]) == [101]          # solo Waldos de diciembre
    assert trunc is False


def test_filtra_solo_mes():
    df, _ = rx.filtrar_facturas(_facturas_demo(), mes=12)
    assert set(df["folio"]) == {101, 103, 104}  # todas las de diciembre


def test_acota_reportes_grandes(monkeypatch):
    monkeypatch.setattr(rx, "MAX_FILAS", 2)
    df, trunc = rx.filtrar_facturas(_facturas_demo(), mes=12)
    assert len(df) == 2 and trunc is True


# ─── Generación del Excel ────────────────────────────────────────────────────

def test_genera_excel_con_encabezados_filas_y_total(tmp_path):
    df, _ = rx.filtrar_facturas(_facturas_demo(), mes=12)
    ruta = rx.generar_excel(df, "Reporte de prueba", ruta=tmp_path / "r.xlsx")
    assert ruta.exists()

    ws = load_workbook(ruta).active
    assert ws["A1"].value == "Reporte de prueba"
    assert [ws.cell(3, c).value for c in range(1, 6)] == \
        ["Folio", "Cliente", "Monto", "Fecha", "Estado"]
    # 3 filas de diciembre + fila TOTAL
    montos = [1500.0, 800.0, 1200.0]
    total_celda = ws.cell(3 + len(montos) + 1, 3).value
    assert total_celda == sum(montos)
    # Estado derivado de fecha_pago
    estados = [ws.cell(r, 5).value for r in range(4, 4 + len(montos))]
    assert "Pagada" in estados and "Pendiente" in estados


# ─── Envío como documento Meta (httpx mockeado) ──────────────────────────────

def _fake_client_factory(capt):
    class _Resp:
        def __init__(self, code, j=None, text=""):
            self.status_code = code
            self._j = j or {}
            self.text = text

        def json(self):
            return self._j

    class _FakeClient:
        def __init__(self, *a, **k):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, *a):
            return False

        async def post(self, url, headers=None, data=None, files=None, json=None):
            if url.endswith("/media"):
                capt["media_type"] = data["type"]
                capt["media_filename"] = files["file"][0]
                return _Resp(200, {"id": "MEDIA_123"})
            capt["doc_payload"] = json
            return _Resp(200, {"messages": [{"id": "wamid.X"}]})

    return _FakeClient


def test_enviar_documento_meta_sube_media_y_manda_document(monkeypatch, tmp_path):
    monkeypatch.setenv("META_ACCESS_TOKEN", "tok")
    monkeypatch.setenv("META_PHONE_NUMBER_ID", "PID")
    capt = {}
    monkeypatch.setattr(webhook.httpx, "AsyncClient", _fake_client_factory(capt))

    ruta = rx.generar_excel(_facturas_demo(), "T", ruta=tmp_path / "rep.xlsx")
    ok = asyncio.run(webhook.enviar_documento_meta(
        "5218681707554", ruta, "reporte.xlsx", caption="hola"))

    assert ok is True
    assert capt["media_filename"] == "reporte.xlsx"
    doc = capt["doc_payload"]
    assert doc["type"] == "document"
    assert doc["document"]["id"] == "MEDIA_123"
    assert doc["document"]["filename"] == "reporte.xlsx"
    assert doc["to"] == "528681707554"   # número MX normalizado (521→52)


def test_enviar_documento_sin_credenciales_retorna_false(monkeypatch, tmp_path):
    monkeypatch.delenv("META_ACCESS_TOKEN", raising=False)
    monkeypatch.delenv("META_PHONE_NUMBER_ID", raising=False)
    ruta = rx.generar_excel(_facturas_demo(), "T", ruta=tmp_path / "rep.xlsx")
    assert asyncio.run(webhook.enviar_documento_meta("521", ruta, "r.xlsx")) is False


# ─── Intercepción SOLO en Meta; Twilio intacto ───────────────────────────────

def _payload_meta(texto):
    return {
        "object": "whatsapp_business_account",
        "entry": [{"changes": [{"value": {
            "messaging_product": "whatsapp",
            "messages": [{"from": "5218681707554", "type": "text",
                          "text": {"body": texto}}],
        }}]}],
    }


def test_meta_reporte_dispara_documento_no_texto(monkeypatch):
    llamado = {"reporte": None, "texto": False}

    async def fake_reporte(numero, solicitud):
        llamado["reporte"] = solicitud

    async def fake_procesar(numero, entrada):
        llamado["texto"] = True
        return "no-deberia"

    monkeypatch.setattr(webhook, "_enviar_reporte_excel_meta", fake_reporte)
    monkeypatch.setattr(webhook, "_procesar_mensaje", fake_procesar)

    client = TestClient(webhook.app)
    r = client.post("/webhook", json=_payload_meta("facturas de waldos de diciembre"))

    assert r.status_code == 200
    assert llamado["reporte"] == {"cliente": "waldos", "mes": 12,
                                  "mes_nombre": "Diciembre", "dias": None, "etiqueta": None}
    assert llamado["texto"] is False   # NO pasó por el flujo de texto


def test_twilio_no_dispara_reporte(monkeypatch):
    """En Twilio el mismo mensaje sigue su flujo de texto (intacto)."""
    monkeypatch.delenv("ENFORCE_TWILIO_SIGNATURE", raising=False)
    monkeypatch.delenv("ENFORCE_WHITELIST", raising=False)
    disparado = {"reporte": False}

    async def fake_reporte(numero, solicitud):
        disparado["reporte"] = True

    async def fake_procesar(numero, entrada):
        return f"texto:{entrada}"

    monkeypatch.setattr(webhook, "_enviar_reporte_excel_meta", fake_reporte)
    monkeypatch.setattr(webhook, "_procesar_mensaje", fake_procesar)

    client = TestClient(webhook.app)
    r = client.post("/webhook",
                    data={"Body": "facturas de waldos de diciembre",
                          "From": "whatsapp:+5218681707554"})

    assert r.status_code == 200
    assert "application/xml" in r.headers["content-type"]   # TwiML (texto)
    assert "texto:facturas de waldos de diciembre" in r.text
    assert disparado["reporte"] is False   # el camino Twilio NO genera documento
